"""Bulk-import projects from CSV or XLSX with auto-categorization into programs.

Accepted columns (case-insensitive, extra columns ignored):
  Required:  project_name
  Optional:  description, program, portfolio, team, status, lead, start_date, end_date

Behavior:
  - If `program` is provided, the project is assigned to that program (created if new).
  - Otherwise, categorize() infers the program from project_name + description + portfolio.
  - Portfolios and teams are auto-created if referenced but missing.
  - All rows are imported transactionally; partial failure returns per-row errors.
"""
import csv
import io
from datetime import datetime
from typing import Any, Iterable
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.hierarchy import Account, Portfolio, Team
from app.models.program import Program, Project
from app.services.categorization import categorize


REQUIRED = {"project_name"}
KNOWN_COLS = {
    "project_name", "description", "program", "portfolio", "team",
    "status", "lead", "start_date", "end_date",
}


def _normalize_headers(fieldnames: Iterable[str]) -> dict[str, str]:
    """Map original header → normalized snake_case lowercase."""
    return {fn: fn.strip().lower().replace(" ", "_") for fn in fieldnames}


def _parse_date(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _read_csv(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    return reader.fieldnames or [], rows


def _read_xlsx(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h or "").strip() for h in next(rows_iter)]
    rows: list[dict] = []
    for row in rows_iter:
        rec = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        if any(v not in (None, "") for v in rec.values()):
            rows.append(rec)
    return headers, rows


async def _get_or_create_account(db: AsyncSession) -> Account:
    res = await db.execute(select(Account).limit(1))
    account = res.scalar_one_or_none()
    if account is None:
        account = Account(name="AT&T")
        db.add(account)
        await db.flush()
    return account


async def _get_or_create_portfolio(db: AsyncSession, account: Account, name: str) -> Portfolio:
    name = name.strip()
    res = await db.execute(select(Portfolio).where(Portfolio.name == name))
    p = res.scalar_one_or_none()
    if p is None:
        p = Portfolio(account_id=account.id, name=name)
        db.add(p)
        await db.flush()
    return p


async def _get_or_create_team(db: AsyncSession, name: str, portfolio: Portfolio | None) -> Team:
    name = name.strip()
    res = await db.execute(select(Team).where(Team.name == name))
    t = res.scalar_one_or_none()
    if t is None:
        t = Team(name=name, portfolio_id=portfolio.id if portfolio else None)
        db.add(t)
        await db.flush()
    return t


async def _get_or_create_program(db: AsyncSession, account: Account, name: str) -> Program:
    name = name.strip()
    res = await db.execute(select(Program).where(Program.name == name))
    p = res.scalar_one_or_none()
    if p is None:
        p = Program(account_id=account.id, name=name, status="planning")
        db.add(p)
        await db.flush()
    return p


async def import_projects(filename: str, file_bytes: bytes, db: AsyncSession) -> dict:
    if filename.lower().endswith(".csv"):
        headers, rows = _read_csv(file_bytes)
    elif filename.lower().endswith((".xlsx", ".xlsm")):
        headers, rows = _read_xlsx(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: {filename}")

    header_map = _normalize_headers(headers)
    normalized_set = set(header_map.values())
    missing = REQUIRED - normalized_set
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}")

    def get(row: dict, col: str) -> str:
        for orig, norm in header_map.items():
            if norm == col:
                v = row.get(orig)
                return "" if v is None else str(v).strip()
        return ""

    account = await _get_or_create_account(db)

    imported: list[dict] = []
    errors: list[dict] = []

    for i, row in enumerate(rows, start=2):  # row 1 is the header
        try:
            name = get(row, "project_name")
            if not name:
                continue

            description = get(row, "description")
            explicit_program = get(row, "program")
            portfolio_name = get(row, "portfolio")
            team_name = get(row, "team")
            status = get(row, "status").lower() or "on_track"

            portfolio = (
                await _get_or_create_portfolio(db, account, portfolio_name)
                if portfolio_name else None
            )
            if team_name:
                await _get_or_create_team(db, team_name, portfolio)

            if explicit_program:
                program_name = explicit_program
                reason = "explicit"
            else:
                program_name, reason = categorize(name, description, portfolio_name)

            program = await _get_or_create_program(db, account, program_name)

            project = Project(program_id=program.id, name=name, status=status)
            db.add(project)
            await db.flush()

            imported.append({
                "row": i,
                "project": name,
                "program": program_name,
                "reason": reason,
            })
        except Exception as e:
            errors.append({"row": i, "data": dict(row), "error": str(e)})

    await db.commit()

    program_counts: dict[str, int] = {}
    for r in imported:
        program_counts[r["program"]] = program_counts.get(r["program"], 0) + 1

    return {
        "imported": len(imported),
        "errors": errors,
        "by_program": program_counts,
        "details": imported,
    }
